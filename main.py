import os,sys,platform,keyboard
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText

# schedule类(排班表)
class Schedule:
    name = []
    date = []
    remark = []
    month = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    month_days = None

    # 初始化
    def __init__(self,workbook):
        self.workbook = workbook
    
    # 格式化文本信息
    def _format(self,str):
        str = str.replace(" ","")
        str = str.replace("\t","")
        str = str.replace("[","")
        str = str.replace("]","")
        return str
    
    # 分割有效信息
    def _split(self,str):
        str = str.strip()
        str = str.replace("\n","")
        if not str:
            return None
        name = str.split("=")
        date = name[1].split(",")
        if ":" in name[1]:
            remark = date[-1].split(":")
            date[-1] = remark[0]
        else:
            remark = ["",""]
        
        return name[0],date,remark[1]
    
    # 读取文件内容
    def loadfile(self,filename):
        with open(filename,'r',encoding='utf-8') as file:
            for line in file.readlines():
                if "#" in line :
                    continue
                data = self._split(line)
                if data != None:
                    self.name.append(data[0])
                    self.date.append(data[1])
                    self.remark.append(data[2])
                
    
    # 导出关键信息txt
    def exportfile(self,filename):
        # 初始化
        wb = self.workbook
        ws = wb.active
        
        # 处理数据
        with open(filename,"w",encoding="utf-8") as file:
            file.write("# 张三=1,2,3,4   :家中有事需要连休\n")
            for row_index in range(4,20):
                if ws.cell(row=row_index,column=1).value == None:
                    break
                self.name.append(ws.cell(row=row_index,column=1).value)
                for column_index in range(2,33):    #处理休息日
                    if ws.cell(row=row_index,column=column_index).value == None:
                        break
                    elif ws.cell(row=row_index,column=column_index).value == "休":
                        self.date.append([])
                        self.date[row_index-4].append(ws.cell(row=3,column=column_index).value)
                remark = ws.cell(row=row_index,column=33).value
                if remark == None:
                    remark = ""
                else:
                    remark = "      :" + remark
                text = f"{self.name[row_index-4]}={self.date[row_index-4]}{remark}\n"
                text = text.replace("[","")
                text = text.replace("]","")
                file.write(text)
        
        # 检测系统并打开txt
        if platform.system() == "Windows":
            os.startfile(filename)
                
    
    # 将txt内容导入排班表
    def importfile(self,filename,to_day,to_night):
        wb = self.workbook
        ws = wb.active
        self.name = []
        self.date = []
        # 格式化
        with open(filename,"r",encoding="utf-8") as file:
            content = file.read()
            content = self._format(content)
        with open(filename,"w",encoding="utf-8") as file:
            file.write(content)
        # 读取txt并写入excel
        with open(filename,"r",encoding="utf-8") as file:
            self.loadfile(filename)
            i = 0
            for row_index, name in enumerate(self.name,start=4):
                i = row_index
                name = self.name[row_index-4]
                ws.cell(row=row_index,column=1,value=hiidle_text(name,name,text_color="000000",bold=True))
                if self.remark[row_index-4] != "":
                    ws.cell(row=row_index,column=33,value=self.remark[row_index-4])
                
                # 修改早/晚/休
                for column_index in range(2,33):
                    target_date = ws.cell(row=3,column=column_index).value
                    # 判断最后本月最后一天
                    if  target_date == "":
                        ws.cell(row=row_index,column=column_index).value = ""
                        continue
                    else:
                        target_date = int(target_date)
                    if target_date > self.month_days:
                        for day,column_index_2 in enumerate(range(column_index,33),start=1):
                            if day <= 20:
                                ws.cell(row=3,column=column_index_2).value = str(day)
                            else:
                                ws.cell(row=2,column=column_index_2).value = ""
                                ws.cell(row=3,column=column_index_2).value = ""
                                ws.cell(row=row_index,column=column_index_2).value = ""
                    
                    if str(target_date) in self.date[row_index-4]:
                        ws.cell(row=row_index,column=column_index).value = hiidle_text("休","休")
                    else:
                        # 修改倒班时间
                        if target_date == to_day or target_date == to_night:
                            ws.cell(row=row_index,column=column_index).value = hiidle_text("倒","倒",text_color="000000")
                        else:   # 修改早，晚班
                            if to_day <=20:
                                if to_night <= target_date <= 31 or target_date < to_day:
                                    ws.cell(row=row_index,column=column_index).value = hiidle_text("晚","晚",text_color="000000")
                                elif to_day < target_date < to_night:
                                    ws.cell(row=row_index,column=column_index).value = hiidle_text("早","早",text_color="000000")
                            else:
                                if to_night < target_date < to_day:
                                    ws.cell(row=row_index,column=column_index).value = hiidle_text("晚","晚",text_color="000000")
                                elif 1 <= target_date < to_night or to_day < target_date:
                                    ws.cell(row=row_index,column=column_index).value = hiidle_text("早","早",text_color="000000")
                        
                        
            # 清空剩余表格
            for index in range(i+1,20):
                for column_index in range(1,34):
                    ws.cell(row=index,column=column_index).value = ""



# 获取年月
def fetch_year_month():
    year_now = datetime.now().year
    month_now = datetime.now().month
    
    while True:
        try:
            year = input(f"请输入当前年份（默认为{year_now}年）：")
            if year == "":
                year = year_now
                break
            year = int(year)
            if 0<=year<=3000:
                break
            else:
                print("请输入正确的年份！")

        except ValueError:
            print("请输入正确的年份！")

    while True:
        try:
            month = input(f"请输入当前月份（默认为{month_now}月）：")
            if month == "":
                month = month_now
                break
            month = int(month)
            if 1<=month<=12:
                break
            else:
                print("请输入正确的月份！")

        except ValueError:
            print("请输入正确的月份！")

    if month == 12:
        title = (f"{year}年{month}月21日至{year+1}年1月20日首钢一期安全员考勤排班表")
    else:
        title = (f"{year}年{month}月21日至{year}年{month+1}月20日首钢一期安全员考勤排班表")
    return title,year,month

# 获取夜转白日期
def fetch_night_to_day():
    while True:
        try:
            to_day = input("请输入一个夜转白的日期（默认为1）：")
            if to_day == "":
                to_day = 1
                break
            to_day = int(to_day)
            if to_day > sch.month_days:
                print("不存在这个日期！")
                continue
            break
        except ValueError:
            pass
    return to_day

# 获取白转夜日期
def fetch_day_to_night():
    while True:
        try:
            to_night = input("请输入一个白转夜的日期（默认为15）：")
            if to_night == "":
                to_night = 15
                break
            to_night = int(to_night)
            if to_night > sch.month_days:
                print("不存在这个日期！")
                continue
            break
        except ValueError:
            pass
    return to_night

# 高亮关键字
def hiidle_text(text,red_text,text_color="C00000",size=17,bold=False,font_style="黑体"):
    rich_text = CellRichText()
    if text == red_text:
        rich_text.append(TextBlock(InlineFont(color=text_color, sz=size, rFont=font_style, b=bold), text))
        return rich_text
    start_index = text.find(red_text)
    end_index = start_index + len(red_text)
    rich_text.append(TextBlock(InlineFont(color="000000",sz=size,rFont=font_style, b=bold), text[:start_index]))
    rich_text.append(TextBlock(InlineFont(color=text_color,sz=size,rFont=font_style), red_text))
    rich_text.append(TextBlock(InlineFont(color="000000",sz=size,rFont=font_style, b=bold), text[end_index:]))
    return rich_text


if __name__ == "__main__":
    if not os.path.isfile("排班.xlsx"):
        input("缺少文件 排班.xlsx ！\n请按任意键继续...")
        sys.exit(0)
    # 初始化
    wb = load_workbook("排班.xlsx")
    ws = wb.active
    sch = Schedule(wb)
    
    # 修改表头
    title_year_month = fetch_year_month()
    title = title_year_month[0]
    red_text = "首钢一期安全员"
    title = hiidle_text(title,red_text,size=20,font_style="微软雅黑")
    ws['A1'].value = title
    # 计算月天数
    year = title_year_month[1]
    month = title_year_month[2]
    if year % 400 == 0 or (year % 100 != 0 and year % 4 == 0):
        sch.month[1] = 29
    sch.month_days = sch.month[month-1]

    # 获取列表关键信息导出txt
    sch.exportfile("content.txt")
    input("请打开 content.txt 并修改内容，完成后按回车键继续")
    # 获取转班日期
    to_day = fetch_night_to_day()
    to_night = fetch_day_to_night()
    
    sch.importfile("content.txt",to_day,to_night)
    remark = hiidle_text("  备注：“√”表示正常上班；   “早”表示早班；  “中”表示中班；“晚”表示晚班；“休”表示休息；“事”表示事假；“年”表示年假；“病”表示病假；“节”表示法定节假；“值”表示值班；“倒”表示倒班","“休”",bold=True)
    ws['A20'].value = remark

    wb.save("output.xlsx")
    print("已导出排班表 output.xlsx")